from openpyxl import load_workbook
workbook = load_workbook(filename="Análises.xlsx")
workbook.sheetnames
sheet=workbook.active


'''
totalcells = 0
total = 0
currentCell = sheet["B26"]
while(total<100):
    totalcells +=1
    currentCell = sheet.cell(row=26, column=currentCell.column+2)
    print(currentCell.value)
    
    total += int(currentCell.value)
    
print(totalcells/total)
'''
print(sheet.cell(row=26,column=2).value)
